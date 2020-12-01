import pandas as pd
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow, Flow
from google.auth.transport.requests import Request
import os
import pickle
import sys
import calendar
import re
from openpyxl import load_workbook
from datetime import datetime
from PyQt5 import (QtGui, QtCore, uic)
from PyQt5.QtWidgets import (QMainWindow, QApplication, QTableWidgetItem, QInputDialog, QErrorMessage, QFileDialog)


# Modify the paths for when the script is being run in a frozen state (i.e. as an EXE)
if getattr(sys, 'frozen', False):
    application_path = sys.executable
    generator_ui_file = 'qt\\report_generator.ui'
    icons_path = 'qt\\icons'
else:
    application_path = os.path.dirname(os.path.abspath(__file__))
    generator_ui_file = os.path.join(application_path, 'qt\\report_generator.ui')
    icons_path = os.path.join(application_path, "qt\\icons")

# Load Qt ui file into a class
generator_ui, _ = uic.loadUiType(generator_ui_file)

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

# here enter the id of your google sheet
SAMPLE_SPREADSHEET_ID_input = open('..//sheet.id', 'r').read()
SAMPLE_RANGE_NAME = 'A3:Q368'  # 1 year of rows


def get_sheet_df(sheet_id):
    # global values_input, service
    print("Retrieving sheet data")
    creds = None
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                '..//credentials.json', SCOPES)  # here enter the name of your downloaded JSON file
            creds = flow.run_local_server(port=0)
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    service = build('sheets', 'v4', credentials=creds)

    # Call the Sheets API
    sheet = service.spreadsheets()
    result_input = sheet.values().get(spreadsheetId=sheet_id,
                                      range=SAMPLE_RANGE_NAME).execute()
    values_input = result_input.get('values', [])

    if not values_input:
        print('No data found.')

    df = pd.DataFrame(values_input[1:], columns=values_input[0])
    return df


class ReportGenerator(QMainWindow, generator_ui):

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.err_msg = QErrorMessage()
        self.table.setColumnWidth(0, 200)

        self.df = pd.DataFrame()

        # Add every year from 2020 to the current year as drop downs
        for year in range(2020, datetime.today().year + 1):
            self.year_cbox.addItem(str(year))
        self.year_cbox.setCurrentText('2020')

        # Signals
        self.year_cbox.currentIndexChanged.connect(self.update_table)
        self.month_cbox.currentIndexChanged.connect(self.update_table)
        self.update_table_btn.clicked.connect(self.update_data)
        self.update_table_btn.clicked.connect(self.update_table)
        self.actionGoogle_Sheets_ID.triggered.connect(self.update_sheets_id)
        self.generate_files_btn.clicked.connect(self.generate_files)

        self.update_data()
        self.update_table()

    def update_sheets_id(self):
        sheet_id, ok_pressed = QInputDialog.getText(self, 'Google Sheets ID', 'Enter Google Sheets ID:')

        if ok_pressed:
            id_file = open("..//sheet.id", "w+")
            id_file.write(sheet_id)
            print(f"New sheet ID: {open('..//sheet.id', 'r').read()}")
            id_file.close()

    def update_data(self):

        print("Updating sheet data")
        try:
            self.df = get_sheet_df(open('..//sheet.id', 'r').read())
        except Exception:
            self.err_msg.showMessage(f"Error retrieving Google Sheets data. "
                                     f"Please make sure the Google Sheets ID is correct.")
            self.df = pd.DataFrame()

    def format_df(self, df, month, year):
        # Remove entries with no comments
        df = df[df['Comments'].astype(bool)]
        # Filter df to only include IRAP
        df = df[df['Comments'].str.contains('IRAP')].reset_index(drop=True)
        # Only keep relevant columns
        df = df.loc[:, ['Date', 'Comments']]
        # Format the date column
        df.Date = df.Date.map(lambda x: datetime.strptime(x, r'%a, %b %d %Y'))
        # Filter df to only include selected month and year
        df = df[df.Date.map(lambda x: x.month == month and x.year == year)]
        return df

    def update_table(self):
        print("Updating table")
        self.table.clearContents()
        if self.df.empty:
            print(f"DataFrame is empty")
            return

        # self.table.setRowCount(0)

        month = self.month_cbox.currentIndex() + 1
        year = int(self.year_cbox.currentText())

        # Add each day of the month as a row
        dates = []
        num_days = calendar.monthrange(year, month)[1]
        self.table.setRowCount(num_days)
        for day in range(1, num_days + 1):
            date = datetime(year, month, day)
            dates.append(date)
            date_str = date.strftime(r'%B %d, %Y (%A)')
            item = QTableWidgetItem(date_str)
            item.setFlags(item.flags() ^ QtCore.Qt.ItemIsEditable)
            self.table.setItem(day - 1, 0, item)

            # # Color the background of weekends gray
            # if date.weekday() in [5, 6]:
            #     for j in range(self.table.columnCount()):
            #         self.table.item(day - 1, j).setBackground(QtGui.QColor(125, 125, 125))
            # else:
            #     for j in range(self.table.columnCount()):
            #         self.table.item(day - 1, j).setBackground(QtGui.QColor(255, 255, 255))

        df = self.format_df(self.df, month, year)
        if not df.empty:
            # Add the description and IRAP hours
            for ind, row in df.iterrows():
                # Split by periods into lists.
                comments = row.Comments.split('\n')

                for comment in comments:
                    irap_re = re.search(r'IRAP:(.*)[({[](.*)[)\]}]\.', comment, re.IGNORECASE)
                    if irap_re:
                        comment = irap_re.group(1).strip()
                        hours = irap_re.group(2).strip()

                        hours_item = QTableWidgetItem(hours)
                        hours_item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
                        self.table.setItem(row.Date.day - 1, 3, QTableWidgetItem(f"{comment}."))
                        self.table.setItem(row.Date.day - 1, 1, hours_item)
                        # continue

        self.table.resizeRowsToContents()

    def generate_files(self):

        def table_to_dataframe():
            number_of_rows = self.table.rowCount()
            number_of_columns = self.table.columnCount()

            df = pd.DataFrame(
                columns=['Date', 'Hours', 'Task_number', 'Description'],  # Fill columnets
                index=range(number_of_rows)  # Fill rows
            )

            for i in range(number_of_rows):
                for j in range(number_of_columns):
                    item = self.table.item(i, j)
                    if item:
                        df.iloc[i, j] = item.text()
                    else:
                        df.iloc[i, j] = None

            return df

        print(f"Generating files")

        month = self.month_cbox.currentText()
        year = self.year_cbox.currentText()

        folder = QFileDialog.getExistingDirectory(self, "Selected Output Folder")
        if not folder:
            return

        data = table_to_dataframe()
        data.Date = data.Date.map(lambda x: datetime.strptime(x, r'%B %d, %Y (%A)'))
        data.dropna(subset=['Hours'], axis=0, inplace=True)

        cells_dict = {
            '1': (18, 4), '2': (19, 4), '3': (20, 4), '4': (21, 4), '5': (22, 4), '6': (23, 4), '7': (24, 4),
            '8': (25, 4), '9': (18, 7), '10': (19, 7), '11': (20, 7), '12': (21, 7), '13': (22, 7), '14': (23, 7),
            '15': (24, 7), '16': (25, 7), '17': (18, 10), '18': (19, 10), '19': (20, 10), '20': (21, 10),
            '21': (22, 10), '22': (23, 10), '23': (24, 10), '24': (25, 10), '25': (18, 13), '26': (19, 13),
            '27': (20, 13), '28': (21, 13), '29': (22, 13), '30': (23, 13), '31': (24, 13)
        }

        excel_file = load_workbook(r'../timesheet_template.xlsx')
        ws = excel_file['Sheet1']

        # Fill the hours
        for ind, row in data.iterrows():
            cell_tup = cells_dict[str(row.Date.day)]
            cell = ws.cell(cell_tup[0], cell_tup[1])

            hours = row.Hours
            cell.value = hours

        # Add the month and year
        ws.cell(10, 11).value = month
        ws.cell(10, 14).value = year

        excel_file.save(f"{folder}\\{month} {year} IRAP Time Sheet.xlsx")


if __name__ == '__main__':
    app = QApplication(sys.argv)

    lc = ReportGenerator()
    lc.show()

    # lc.generate_files()

    app.exec_()
